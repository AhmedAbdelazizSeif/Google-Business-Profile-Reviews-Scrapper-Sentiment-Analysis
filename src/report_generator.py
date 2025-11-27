"""
Excel Report Generator Module

This module provides functionality to create beautifully styled Excel reports
with automatic brand color extraction from logos.
"""

import os
import pandas as pd
from datetime import datetime
from typing import Dict, Tuple, List, Optional
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image


class ReportGenerator:
    """
    Generates professional Excel reports with branded styling.
    
    Features:
    - Automatic color palette extraction from logo
    - Multi-sheet workbooks
    - Logo integration
    - Alternating row colors
    - Frozen header panes
    - Auto-adjusted column widths
    """
    
    # Default color palette (fallback)
    DEFAULT_COLORS = {
        'primary': 'FF6B35',      # Orange
        'secondary': '004E89',    # Blue
        'accent1': 'F7B801',      # Yellow
        'accent2': '1B998B',      # Teal
        'light_gray': 'F4F4F4',
        'medium_gray': 'E0E0E0',
        'dark_gray': '666666',
        'white': 'FFFFFF',
        'text_dark': '2C2C2C',
    }
    
    def __init__(self, logo_path: Optional[str] = None):
        """
        Initialize report generator.
        
        Args:
            logo_path: Path to brand logo image (optional)
        """
        self.logo_path = logo_path
        self.colors = self._extract_or_use_default_colors()
    
    def _extract_or_use_default_colors(self) -> Dict[str, str]:
        """Extract colors from logo or use defaults."""
        if self.logo_path and os.path.exists(self.logo_path):
            extracted = self.extract_color_palette(self.logo_path)
            if extracted and len(extracted) >= 4:
                return {
                    'primary': extracted[0],
                    'secondary': extracted[1],
                    'accent1': extracted[2],
                    'accent2': extracted[3],
                    **{k: v for k, v in self.DEFAULT_COLORS.items() 
                       if k not in ['primary', 'secondary', 'accent1', 'accent2']}
                }
        return self.DEFAULT_COLORS.copy()
    
    @staticmethod
    def extract_color_palette(
        image_path: str, 
        num_colors: int = 8
    ) -> Optional[List[str]]:
        """
        Extract dominant colors from an image.
        
        Args:
            image_path: Path to image file
            num_colors: Number of colors to extract
            
        Returns:
            List of hex color codes or None if extraction fails
        """
        try:
            img = Image.open(image_path)
            img = img.convert('RGB')
            img = img.resize((150, 150))
            
            pixels = list(img.getdata())
            
            # Filter out very light and very dark colors
            filtered_pixels = [p for p in pixels if 50 < sum(p) < 700]
            
            # Count color frequency
            color_counter = Counter(filtered_pixels)
            most_common = color_counter.most_common(num_colors * 3)
            
            # Convert to hex and remove duplicates
            unique_colors = []
            for color, _ in most_common:
                hex_color = '{:02X}{:02X}{:02X}'.format(*color)
                if hex_color not in unique_colors and hex_color not in ['FFFFFF', '000000']:
                    unique_colors.append(hex_color)
                if len(unique_colors) >= num_colors:
                    break
            
            return unique_colors
        except Exception as e:
            print(f"Could not extract colors from logo: {e}")
            return None
    
    def style_worksheet(
        self,
        ws,
        df: pd.DataFrame,
        sheet_title: str,
        color_scheme: str = 'primary'
    ):
        """
        Apply professional styling to a worksheet.
        
        Args:
            ws: Worksheet object
            df: DataFrame to write
            sheet_title: Title for the sheet
            color_scheme: 'primary' or 'secondary' color theme
        """
        # Choose colors
        header_color = (self.colors['primary'] if color_scheme == 'primary' 
                       else self.colors['secondary'])
        title_bg_color = (self.colors['secondary'] if color_scheme == 'primary' 
                         else self.colors['primary'])
        
        current_row = 1
        
        # Row 1: Add logo if exists
        if self.logo_path and os.path.exists(self.logo_path):
            try:
                img = XLImage(self.logo_path)
                aspect_ratio = img.width / img.height
                img.height = 60
                img.width = int(60 * aspect_ratio)
                ws.add_image(img, 'A1')
                ws.row_dimensions[1].height = 45
                current_row = 2
            except Exception as e:
                print(f"Could not add logo: {e}")
        
        # Row 2: Title row
        title_row = current_row
        ws.merge_cells(f'A{title_row}:{get_column_letter(len(df.columns))}{title_row}')
        title_cell = ws[f'A{title_row}']
        title_cell.value = sheet_title
        title_cell.font = Font(name='Segoe UI', size=16, bold=True, 
                              color=self.colors['white'])
        title_cell.fill = PatternFill(start_color=title_bg_color, 
                                      end_color=title_bg_color, fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[title_row].height = 35
        
        # Row 3+: Headers
        header_row = title_row + 1
        
        for col_num, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = str(column_name)
            cell.font = Font(name='Segoe UI', size=11, bold=True, 
                           color=self.colors['white'])
            cell.fill = PatternFill(start_color=header_color, 
                                   end_color=header_color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', 
                                      wrap_text=True)
            cell.border = Border(bottom=Side(style='medium', 
                                            color=self.colors['dark_gray']))
        
        # Data rows with alternating colors
        for row_num, row_data in enumerate(df.values, header_row + 1):
            is_even = (row_num - header_row) % 2 == 0
            row_fill_color = (self.colors['white'] if is_even 
                            else self.colors['light_gray'])
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.font = Font(name='Segoe UI', size=10, 
                               color=self.colors['text_dark'])
                cell.fill = PatternFill(start_color=row_fill_color, 
                                       end_color=row_fill_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center', 
                                          wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin', color=self.colors['medium_gray']),
                    right=Side(style='thin', color=self.colors['medium_gray']),
                    bottom=Side(style='thin', color=self.colors['medium_gray'])
                )
        
        # Auto-adjust column widths
        for col_num, column_name in enumerate(df.columns, 1):
            max_length = len(str(column_name))
            for row_num in range(header_row + 1, header_row + len(df) + 1):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width
        
        # Freeze panes
        ws.freeze_panes = ws[f'A{header_row + 1}']
    
    def create_workbook(
        self,
        dataframes_dict: Dict[str, Tuple[pd.DataFrame, str]],
        filename: str
    ) -> str:
        """
        Create a multi-sheet Excel workbook with styled sheets.
        
        Args:
            dataframes_dict: Dict mapping sheet names to (DataFrame, color_scheme) tuples
            filename: Output filename
            
        Returns:
            Path to created file
        """
        wb = Workbook()
        wb.remove(wb.active)
        
        for sheet_name, (df, color_scheme) in dataframes_dict.items():
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel limit
            self.style_worksheet(ws, df, sheet_name, color_scheme)
        
        wb.save(filename)
        print(f"✓ Created workbook: {filename}")
        return filename


# Example usage
if __name__ == "__main__":
    # Sample data
    data = {
        'اسم البائع': ['Ahmed', 'Mohamed', 'Ali'],
        'Positive': [45, 38, 52],
        'Negative': [3, 5, 2],
        'Neutral': [12, 8, 15]
    }
    df = pd.DataFrame(data)
    
    # Create report
    generator = ReportGenerator(logo_path='../assets/logo.png')
    
    workbook_data = {
        '📊 Performance Report': (df, 'primary')
    }
    
    generator.create_workbook(
        workbook_data,
        f'sample_report_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    )
